"""
Monarch Investment — Vendor Aging Property Matcher (Core Logic)
Handles matching, fuzzy scoring, and Excel output generation.
"""

import io
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz, process as fuzz_process
    RAPIDFUZZ_AVAILABLE = True
except ImportError:
    RAPIDFUZZ_AVAILABLE = False


# ---------------------------------------------------------------------------
# Vendor configurations
# Each vendor aging report has its own column layout.
# prop_col     : column containing the property/customer name
# invoice_col  : column containing the invoice number
# sheet        : Excel sheet name to read
# header_row   : 0-based row index for the header (passed to pd.read_excel)
# ffill_prop   : True if the property name only appears on the first row of a
#                group and must be forward-filled down to child rows
# ---------------------------------------------------------------------------
VENDOR_CONFIGS = {
    "Chadwell": {
        "sheet": "Management Company Invoice Summ",
        "header_row": 2,
        "prop_col": "Customer",
        "invoice_col": "Invoice #",
    },
    "Ferguson": {
        "sheet": "Export",
        "header_row": 0,
        "prop_col": "Customer Name",
        "invoice_col": "Invoice ID",
    },
    "HD Supply": {
        "sheet": "Detail",
        "header_row": 0,
        "prop_col": "Name.1",       # 4th column — site-level property name
        "invoice_col": "Invoice",
    },
    "Lowes": {
        "sheet": "transaction",
        "header_row": 4,
        "prop_col": "Account Name",
        "invoice_col": "Transaction Number",
    },
    "Sherwin-Williams": {
        "sheet": "Open_Transcations",
        "header_row": 2,
        "prop_col": "Account Name",
        "invoice_col": "Extended Invoice #",
    },
    "Staples": {
        "sheet": "Invoices",
        "header_row": 0,
        "prop_col": "Bill To Customer",
        "invoice_col": "Number",
    },
    "Lowes Pro": {
        "sheet": "Sheet 1",
        "header_row": 0,
        "prop_col": "Customer Name",
        "invoice_col": "Invoice-suf#",
        "ffill_prop": True,     # property name only on first row of each group
    },
}

VENDOR_NAMES = sorted(VENDOR_CONFIGS.keys())


# ---------------------------------------------------------------------------
# Per-vendor custom lookup persistence
# ---------------------------------------------------------------------------

def _vendor_slug(vendor_name: str) -> str:
    """Convert vendor display name to a safe filename slug."""
    slug = vendor_name.lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug).strip("_")
    return slug


def get_vendor_lookup_filename(vendor_name: str) -> str:
    """Return the JSON filename for a given vendor's lookup table."""
    return f"custom_lookup_{_vendor_slug(vendor_name)}.json"


def load_custom_lookup(path):
    """Load user-confirmed pcodes from a lookup JSON file. Returns {} if missing."""
    try:
        p = Path(path)
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_custom_lookup(data, path):
    """Persist user-confirmed pcodes to a lookup JSON file."""
    with open(Path(path), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

# ---------------------------------------------------------------------------
# CURATED LOOKUP TABLE
# Keys  : vendor property name, normalized (UPPER, stripped, single-spaced)
# Values: pcode, official property name, confidence, needs_review flag
#
# HOW TO ADD NEW ENTRIES:
#   After reviewing a flagged item, add a new line here with "HIGH" confidence
#   and needs_review=False so it matches automatically on every future run.
# ---------------------------------------------------------------------------
CURATED_LOOKUP = {
    "21 SOUTH AT PARKVIEW APARTMENT HOMES": {"pcode": "tsla",  "official_name": "21 South at Parkview",              "confidence": "HIGH",   "needs_review": False},
    "ABBEY COURT APARTMENTS":               {"pcode": "ACIN",  "official_name": "Abbey Court Apartments",            "confidence": "HIGH",   "needs_review": False},
    "ACADIAN POINT APARTMENTS":             {"pcode": "APLA",  "official_name": "Acadian Point Apartments",          "confidence": "HIGH",   "needs_review": False},
    "ADDISON PLACE":                        {"pcode": "apin",  "official_name": "Addison Place",                     "confidence": "HIGH",   "needs_review": False},
    "AFTON OAKS":                           {"pcode": "aola",  "official_name": "Afton Oaks",                        "confidence": "HIGH",   "needs_review": False},
    "ARCADIA GROVE":                        {"pcode": "agmi",  "official_name": "Arcadia Grove",                     "confidence": "HIGH",   "needs_review": False},
    "ARDSLEY RIDGE":                        {"pcode": "ayoh",  "official_name": "Ardsley Ridge",                     "confidence": "HIGH",   "needs_review": False},
    "ASHTON POINTE":                        {"pcode": "attx",  "official_name": "Ashton Pointe",                     "confidence": "HIGH",   "needs_review": False},
    "AVALON PLACE":                         {"pcode": "ANOH",  "official_name": "Avalon Place",                      "confidence": "HIGH",   "needs_review": False},
    "BARRINGTON PLACE AT SOMERSET":         {"pcode": "BPAL",  "official_name": "Barrington Place at Somerset",      "confidence": "HIGH",   "needs_review": False},
    "BAYBERRY PLACE TOWNHOMES":             {"pcode": "BPOH",  "official_name": "Bayberry Place Townhomes",          "confidence": "HIGH",   "needs_review": False},
    "BEACON HILL":                          {"pcode": "BHAR",  "official_name": "Beacon Hill",                       "confidence": "HIGH",   "needs_review": False},
    "BLOOMFIELD TOWNHOMES":                 {"pcode": "blmi",  "official_name": "Bloomfield Townhomes",              "confidence": "HIGH",   "needs_review": False},
    "BLUEGRASS VILLAS":                     {"pcode": "BVKY",  "official_name": "Bluegrass Villas",                  "confidence": "HIGH",   "needs_review": False},
    "BRONCO CLUB":                          {"pcode": "bcmi",  "official_name": "Bronco Club",                       "confidence": "HIGH",   "needs_review": False},
    "BURWICK FARMS":                        {"pcode": "bfmi",  "official_name": "Burwick Farms Apartments",          "confidence": "HIGH",   "needs_review": False},
    "CARRIAGE HILL APARTMENTS":             {"pcode": "choh",  "official_name": "Carriage Hill",                     "confidence": "HIGH",   "needs_review": False},
    "CEDAR TRACE":                          {"pcode": "ctmo",  "official_name": "Cedar Trace",                       "confidence": "HIGH",   "needs_review": False},
    "CEDARWOOD VILLAGE":                    {"pcode": "cwoh",  "official_name": "Cedarwood Village",                 "confidence": "HIGH",   "needs_review": False},
    "CENTRAL SQUARE APARTMENTS":            {"pcode": "csoh",  "official_name": "Central Square Apartments",         "confidence": "HIGH",   "needs_review": False},
    "CHATEAU RIVIERA":                      {"pcode": "ctmi",  "official_name": "Chateau Riviera",                   "confidence": "HIGH",   "needs_review": False},
    "CHELSEA PARK":                         {"pcode": "cpmi",  "official_name": "Chelsea Park Apartments",           "confidence": "HIGH",   "needs_review": False},
    "CLAIR COMMONS":                        {"pcode": "cmoh",  "official_name": "Clair Commons",                     "confidence": "HIGH",   "needs_review": False},
    "COLONY VILLAGE APTS":                  {"pcode": "cvnc",  "official_name": "Colony Village",                    "confidence": "HIGH",   "needs_review": False},
    "COUNTRY ESTATES":                      {"pcode": "cene",  "official_name": "Country Estates Townhomes",         "confidence": "HIGH",   "needs_review": False},
    "CROSS CREEK":                          {"pcode": "CRKS",  "official_name": "Cross Creek Apartments",            "confidence": "HIGH",   "needs_review": False},
    "CROWN COLONY APARTMENTS":              {"pcode": "ccks",  "official_name": "Crown Colony",                      "confidence": "HIGH",   "needs_review": False},
    "CROWN POINTE":                         {"pcode": "crmi",  "official_name": "Crown Pointe Apartments",           "confidence": "HIGH",   "needs_review": False},
    "DOGWOOD APARTMENTS":                   {"pcode": "DWOH",  "official_name": "Dogwood Apartment",                 "confidence": "HIGH",   "needs_review": False},
    "DOVE CREEK APARTMENTS":                {"pcode": "dcla",  "official_name": "Dove Creek Apartments",             "confidence": "HIGH",   "needs_review": False},
    "EAGLE RIDGE APARTMENTS":               {"pcode": "eroh",  "official_name": "Eagle Ridge Apartments",            "confidence": "HIGH",   "needs_review": False},
    "EASTWOOD ARMS":                        {"pcode": "ewoh",  "official_name": "Eastwood Arms",                     "confidence": "HIGH",   "needs_review": False},
    "EDGE AT ARLINGTON":                    {"pcode": "EAOH",  "official_name": "Edge at Arlington",                 "confidence": "HIGH",   "needs_review": False},
    "FAIRFIELD":                            {"pcode": "ffmi",  "official_name": "Fairfield Apartments and Condos",   "confidence": "LOW",    "needs_review": True},
    "FOREST HILLS":                         {"pcode": "fhnc",  "official_name": "Forest Hills",                      "confidence": "HIGH",   "needs_review": False},
    "FOREST WOODS":                         {"pcode": "fwmo",  "official_name": "Forest Woods",                      "confidence": "HIGH",   "needs_review": False},
    "GEORGETOWNE APARTMENTS":               {"pcode": "gtne",  "official_name": "Georgetowne Apartment Homes",       "confidence": "HIGH",   "needs_review": False},
    "GREENWAY CHASE APARTMENTS":            {"pcode": "GCMO",  "official_name": "Greenway Chase",                    "confidence": "HIGH",   "needs_review": False},
    "HAMILTON TRACE":                       {"pcode": "htmi",  "official_name": "Hamilton Trace Apartments",         "confidence": "HIGH",   "needs_review": False},
    "HEARTH  HOLLOW APARTMENTS":            {"pcode": "hhks",  "official_name": "Hearth Hollow",                     "confidence": "HIGH",   "needs_review": False},
    "HEARTH HOLLOW APARTMENTS":             {"pcode": "hhks",  "official_name": "Hearth Hollow",                     "confidence": "HIGH",   "needs_review": False},
    "HIDDEN TREE APARTMENTS":               {"pcode": "HEMI",  "official_name": "Hidden Tree",                       "confidence": "HIGH",   "needs_review": False},
    "HUNTER'S RIDGE":                       {"pcode": "hroh",  "official_name": "Hunters Ridge",                     "confidence": "MEDIUM", "needs_review": True},
    "HUNTERS RIDGE":                        {"pcode": "hroh",  "official_name": "Hunters Ridge",                     "confidence": "MEDIUM", "needs_review": True},
    "HUNTERS RIDGE APARTMENTS":             {"pcode": "hrmo",  "official_name": "Hunters Ridge Apartments",          "confidence": "HIGH",   "needs_review": False},
    "HUNTLEY RIDGE NEW ALBANY":             {"pcode": "hain",  "official_name": "Huntley Ridge New Albany",          "confidence": "HIGH",   "needs_review": False},
    "HUNTLEY RIDGE OLDE TOWNE":             {"pcode": "otin",  "official_name": "Olde Towne Village",                "confidence": "LOW",    "needs_review": True},
    "HUNTLEY RIDGE TOWNHOMES":              {"pcode": "htin",  "official_name": "Huntley Ridge Townhomes",           "confidence": "HIGH",   "needs_review": False},
    "ICON LOUISVILLE":                      {"pcode": "icky",  "official_name": "Icon",                              "confidence": "HIGH",   "needs_review": False},
    "INDIAN SPRINGS APARTMENTS":            {"pcode": "isin",  "official_name": "Indian Springs",                    "confidence": "HIGH",   "needs_review": False},
    "INDIAN WOODS TOWNHOMES":               {"pcode": "iwin",  "official_name": "Indian Woods",                      "confidence": "HIGH",   "needs_review": False},
    "LAKEWOOD":                             {"pcode": "lwmi",  "official_name": "Lakewood Apartments",               "confidence": "HIGH",   "needs_review": False},
    "LAKOTA LAKE":                          {"pcode": "lloh",  "official_name": "Lakota Lake Apartments",            "confidence": "HIGH",   "needs_review": False},
    "LAMBERTON LAKES APTS":                 {"pcode": "LAMI",  "official_name": "Lamberton Lake",                    "confidence": "HIGH",   "needs_review": False},
    "LANSING TOWER APARTMENTS":             {"pcode": "LTMI",  "official_name": "Lansing Towers",                    "confidence": "HIGH",   "needs_review": False},
    "MADISON GROVE":                        {"pcode": "troh",  "official_name": "Madison Grove Townhomes",           "confidence": "HIGH",   "needs_review": False},
    "MCMILLEN WOODS":                       {"pcode": "mwoh",  "official_name": "McMillen Woods Apartments",         "confidence": "HIGH",   "needs_review": False},
    "MEADOW LARK APARTMENTS, LLC":          {"pcode": "ml",    "official_name": "Meadow Lark Apartments",            "confidence": "HIGH",   "needs_review": False},
    "MILLER WEST":                          {"pcode": "MWMI",  "official_name": "Miller West",                       "confidence": "HIGH",   "needs_review": False},
    "MT. CARMEL":                           {"pcode": "mcks",  "official_name": "Mt. Carmel Village Apartments",     "confidence": "HIGH",   "needs_review": False},
    "NANTUCKET GARDENS APARTMENTS":         {"pcode": "ngmo",  "official_name": "Nantucket Gardens",                 "confidence": "HIGH",   "needs_review": False},
    "NEMOKE TRAILS":                        {"pcode": "ntmi",  "official_name": "Nemoke Trails Apartments",          "confidence": "HIGH",   "needs_review": False},
    "NORTH PARK APARTMENTS":                {"pcode": "npin",  "official_name": "North Park",                        "confidence": "HIGH",   "needs_review": False},
    "OAKWOOD VILLAGE":                      {"pcode": "OVWV",  "official_name": "Oakwood Village",                   "confidence": "HIGH",   "needs_review": False},
    "PADDOCK VILLAGE":                      {"pcode": "pvmo",  "official_name": "Paddock Village",                   "confidence": "HIGH",   "needs_review": False},
    "PARCSTONE APARTMENTS":                 {"pcode": "PSNC",  "official_name": "Parcstone Apartments",              "confidence": "HIGH",   "needs_review": False},
    "PAVILION LAKES":                       {"pcode": "plin",  "official_name": "Pavilion Lakes",                    "confidence": "HIGH",   "needs_review": False},
    "PECAN GROVE":                          {"pcode": "pgla",  "official_name": "Pecan Grove Apartments",            "confidence": "HIGH",   "needs_review": False},
    "PINE GROVE":                           {"pcode": "PGSC",  "official_name": "Pine Grove",                        "confidence": "HIGH",   "needs_review": False},
    "PINE RUN TOWNHOMES":                   {"pcode": "proh",  "official_name": "Pine Run Townhomes",                 "confidence": "HIGH",   "needs_review": False},
    "PRESTON OAKS APTS":                    {"pcode": "poky",  "official_name": "Preston Oaks",                      "confidence": "HIGH",   "needs_review": False},
    "QUAIL HOLLOW":                         {"pcode": "qhsc",  "official_name": "Quail Hollow",                      "confidence": "HIGH",   "needs_review": False},
    "RETREAT AT SEVEN TRAILS":              {"pcode": "stmo",  "official_name": "The Retreat at Seven Trails",       "confidence": "HIGH",   "needs_review": False},
    "RETREAT AT WOODRIDGE":                 {"pcode": "wrks",  "official_name": "The Retreat at Woodridge",          "confidence": "HIGH",   "needs_review": False},
    "RIDGE AT CHESTNUT":                    {"pcode": "trmo",  "official_name": "The Ridge at Chestnut",             "confidence": "HIGH",   "needs_review": False},
    "RIVER CHASE":                          {"pcode": "rcmo",  "official_name": "River Chase Apartments",            "confidence": "HIGH",   "needs_review": False},
    "RIVERWALK":                            {"pcode": "rwmi",  "official_name": "Riverwalk Apartments",              "confidence": "HIGH",   "needs_review": False},
    "ROLLING PINES":                        {"pcode": "rpmi",  "official_name": "Rolling Pines Apartments",          "confidence": "HIGH",   "needs_review": False},
    "ROSEMOORE":                            {"pcode": "rmil",  "official_name": "Rosemoore Portfolio",               "confidence": "HIGH",   "needs_review": False},
    "SHERWOOD APARTMENTS":                  {"pcode": "swks",  "official_name": "Sherwood",                          "confidence": "HIGH",   "needs_review": False},
    "SHORES OF ROOSEVELT PARK":             {"pcode": "srmi",  "official_name": "The Shores of Roosevelt Park",      "confidence": "HIGH",   "needs_review": False},
    "SILVER LAKE HILLS":                    {"pcode": "svmi",  "official_name": "Silver Lake Hills Apartments",      "confidence": "HIGH",   "needs_review": False},
    "SOUTHGATE":                            {"pcode": "sgnc",  "official_name": "Southgate Apartments",              "confidence": "HIGH",   "needs_review": False},
    "SOUTHWOODS APARTMENTS":                {"pcode": "swmo",  "official_name": "Southwoods",                        "confidence": "HIGH",   "needs_review": False},
    "SPRINGCREEK":                          {"pcode": "scks",  "official_name": "Springcreek",                       "confidence": "HIGH",   "needs_review": False},
    "SUGAR PINES":                          {"pcode": "sgmo",  "official_name": "Sugar Pines Apartments",            "confidence": "HIGH",   "needs_review": False},
    "SUMMER BROOK":                         {"pcode": "sbtx",  "official_name": "Summer Brook",                      "confidence": "HIGH",   "needs_review": False},
    "SUMMERHOUSE SQUARE":                   {"pcode": "ssoh",  "official_name": "Summerhouse Square",                "confidence": "HIGH",   "needs_review": False},
    "SUMMIT ON THE LAKE":                   {"pcode": "sltx",  "official_name": "Summit on the Lake Apartments",     "confidence": "HIGH",   "needs_review": False},
    "SUNNYDALE ESTATES":                    {"pcode": "seoh",  "official_name": "Sunnydale Estates",                 "confidence": "HIGH",   "needs_review": False},
    "TAYLOR SQUARE":                        {"pcode": "tsoh",  "official_name": "Taylor Square",                     "confidence": "HIGH",   "needs_review": False},
    "THE BERKLEY":                          {"pcode": "tbar",  "official_name": "The Berkley Apartments",            "confidence": "HIGH",   "needs_review": False},
    "THE BRICKYARD APARTMENTS":             {"pcode": "tbin",  "official_name": "The Brickyard",                     "confidence": "HIGH",   "needs_review": False},
    "THE BROOKEVILLE":                      {"pcode": "tboh",  "official_name": "The Brookeville Apartments",        "confidence": "HIGH",   "needs_review": False},
    "THE COLE":                             {"pcode": "clnc",  "official_name": "The Cole Apartments",               "confidence": "HIGH",   "needs_review": False},
    "THE DISTRICT":                         {"pcode": "TDMO",  "official_name": "The District",                      "confidence": "HIGH",   "needs_review": False},
    "THE EARL":                             {"pcode": "tefl",  "official_name": "The Earl",                          "confidence": "HIGH",   "needs_review": False},
    "THE HIGHLANDS OF WEST CHESTER":        {"pcode": "thoh",  "official_name": "The Highlands of West Chester",     "confidence": "HIGH",   "needs_review": False},
    "THE HILLS":                            {"pcode": "thmo",  "official_name": "The Hills Apartments",              "confidence": "HIGH",   "needs_review": False},
    "THE JUNCTION AT RAMSEY":               {"pcode": "tjnc",  "official_name": "The Junction at Ramsey & Carver",   "confidence": "HIGH",   "needs_review": False},
    "THE LANDING APARTMENT HOMES":          {"pcode": "tlnc",  "official_name": "The Landing Apartment Homes",       "confidence": "HIGH",   "needs_review": False},
    "THE MADISON":                          {"pcode": "tmnc",  "official_name": "The Madison Apartments",            "confidence": "HIGH",   "needs_review": False},
    "THE OAKS AT PRAIRIE VIEW":             {"pcode": "tomo",  "official_name": "The Oaks At Prairie View",          "confidence": "HIGH",   "needs_review": False},
    "THE PARK":                             {"pcode": "tptx",  "official_name": "The Park",                          "confidence": "MEDIUM", "needs_review": True},
    "THE PARK APARTMENTS HOMES":            {"pcode": "pknc",  "official_name": "The Park Apartment Homes",          "confidence": "HIGH",   "needs_review": False},
    "THE RESIDENZ":                         {"pcode": "rzoh",  "official_name": "The Residenz",                      "confidence": "HIGH",   "needs_review": False},
    "UNION AT COOPER HILL I-IV":            {"pcode": "CHAL",  "official_name": "Cooper Hill",                       "confidence": "LOW",    "needs_review": True},
    "VALLEY STREAM":                        {"pcode": "vsoh",  "official_name": "Valley Stream Apartments",          "confidence": "HIGH",   "needs_review": False},
    "VICINO ON THE LAKE":                   {"pcode": "vlmo",  "official_name": "Vicino on the Lake",                "confidence": "HIGH",   "needs_review": False},
    "VILLAGE ROYALE APARTMENTS":            {"pcode": "vrmo",  "official_name": "Village Royale",                    "confidence": "HIGH",   "needs_review": False},
    "WAKE ROBIN":                           {"pcode": "wroh",  "official_name": "Wake Robin Apartments",             "confidence": "HIGH",   "needs_review": False},
    "WATERFORD PINES":                      {"pcode": "UNKNOWN", "official_name": "NOT FOUND",                       "confidence": "LOW",    "needs_review": True},
    "WATERSTONE PLACE":                     {"pcode": "wsoh",  "official_name": "Waterstone Place",                  "confidence": "HIGH",   "needs_review": False},
    "WEST WINDS":                           {"pcode": "WWSC",  "official_name": "West Winds",                        "confidence": "HIGH",   "needs_review": False},
    "WESTCHESTER VILLAGE":                  {"pcode": "wvmo",  "official_name": "Westchester Village Apartments",    "confidence": "HIGH",   "needs_review": False},
    "WHISPERING TIMBERS":                   {"pcode": "tioh",  "official_name": "Whispering Timbers",                "confidence": "HIGH",   "needs_review": False},
    "WILDEWOOD SOUTH":                      {"pcode": "WSSC",  "official_name": "Wildewood South",                   "confidence": "HIGH",   "needs_review": False},
    "WILLIAMSBURG WAY":                     {"pcode": "wwin",  "official_name": "Williamsburg Way",                  "confidence": "HIGH",   "needs_review": False},
    "WILLOW RUN":                           {"pcode": "wrnc",  "official_name": "Willow Run",                        "confidence": "HIGH",   "needs_review": False},
    "WOODHOLLOW APARTMENTS":                {"pcode": "whmo",  "official_name": "Woodhollow Apartments",             "confidence": "HIGH",   "needs_review": False},
    "WOODS OF POST HOUSE":                  {"pcode": "wptn",  "official_name": "Woods of Post House",               "confidence": "HIGH",   "needs_review": False},
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

import unicodedata


def normalize(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s) if hasattr(unicodedata, "normalize") else s
    s = re.sub(r"\s+", " ", s.strip().upper())
    return s


def load_property_list(path):
    df = pd.read_excel(path, sheet_name="Currently Owned")
    df["Yardi Code"]    = df["Yardi Code"].astype(str).str.strip()
    df["Property Name"] = df["Property Name"].astype(str).str.strip()
    norm_to_pcode = {normalize(row["Property Name"]): (row["Property Name"], row["Yardi Code"])
                     for _, row in df.iterrows()}
    return norm_to_pcode


def fuzzy_match(vendor_norm, norm_to_pcode, threshold):
    if not RAPIDFUZZ_AVAILABLE:
        return "UNKNOWN", "NOT FOUND", 0, True

    prop_names_norm = list(norm_to_pcode.keys())
    best = fuzz_process.extractOne(vendor_norm, prop_names_norm, scorer=fuzz.token_set_ratio)
    if best is None:
        return "UNKNOWN", "NOT FOUND", 0, True

    matched_norm, score, _ = best
    official_name, pcode = norm_to_pcode[matched_norm]
    # Flag both MEDIUM (score < 90) and LOW (score < threshold) for review
    needs_review = score < 90
    return pcode, official_name, score, needs_review


def match_vendor_name(vendor_raw, norm_to_pcode, threshold, custom_lookup=None):
    norm = normalize(vendor_raw)

    # 1. User-confirmed custom lookup (highest priority — overrides everything)
    if custom_lookup and norm in custom_lookup:
        r = custom_lookup[norm].copy()
        r["match_method"] = "Confirmed by User"
        r["match_score"]  = 100
        r["needs_review"] = False
        r["confidence"]   = "HIGH"
        return r

    # 2. Built-in curated lookup
    if norm in CURATED_LOOKUP:
        r = CURATED_LOOKUP[norm].copy()
        r["match_method"] = "Curated Lookup"
        r["match_score"]  = 100
        return r

    if norm in norm_to_pcode:
        official_name, pcode = norm_to_pcode[norm]
        return {"pcode": pcode, "official_name": official_name, "confidence": "HIGH",
                "needs_review": False, "match_method": "Exact Match", "match_score": 100}

    pcode, official_name, score, needs_review = fuzzy_match(norm, norm_to_pcode, threshold)
    confidence = "HIGH" if score >= 90 else ("MEDIUM" if score >= threshold else "LOW")
    return {"pcode": pcode, "official_name": official_name, "confidence": confidence,
            "needs_review": needs_review, "match_method": f"Fuzzy (score={score:.0f})",
            "match_score": score}


def _load_vendor_df(vendor_path, vendor_name):
    """
    Load the vendor aging Excel into a DataFrame using vendor-specific settings.
    Returns (df, prop_col, invoice_col).
    """
    cfg = VENDOR_CONFIGS.get(vendor_name)

    xl = pd.ExcelFile(vendor_path)

    if cfg:
        sheet      = cfg["sheet"] if cfg["sheet"] in xl.sheet_names else xl.sheet_names[0]
        header_row = cfg["header_row"]
        prop_col   = cfg["prop_col"]
        invoice_col = cfg.get("invoice_col")
        ffill_prop  = cfg.get("ffill_prop", False)
    else:
        # Fallback: auto-detect header row by looking for "Invoice #"
        sheet = xl.sheet_names[0]
        df_raw = pd.read_excel(vendor_path, sheet_name=sheet, header=None)
        header_row = 0
        for i, row in df_raw.iterrows():
            if any("Invoice #" in str(v) for v in row.values):
                header_row = i
                break
        ffill_prop  = False
        invoice_col = None
        prop_col    = None

    df = pd.read_excel(vendor_path, sheet_name=sheet, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    # Forward-fill property name for grouped layouts (e.g. Lowes Pro)
    if ffill_prop and prop_col and prop_col in df.columns:
        df[prop_col] = df[prop_col].replace("", pd.NA).ffill()

    # Auto-detect prop_col if not set by config
    if not prop_col:
        prop_col = next((c for c in df.columns if "customer" in c.lower()), None)
        if prop_col is None:
            for c in df.columns:
                sample = df[c].dropna().astype(str)
                if sample.str.len().mean() > 10:
                    prop_col = c
                    break
        if prop_col is None:
            raise ValueError("Could not find the property name / customer column.")

    # Auto-detect invoice_col if not set
    if not invoice_col:
        invoice_col = next((c for c in df.columns if "invoice" in c.lower()), None)

    return df, prop_col, invoice_col


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_matcher(vendor_path, prop_path, output_dest, fuzzy_threshold=75,
                custom_lookup_path=None, vendor_name=None):
    """
    vendor_path        : path or file-like for the vendor aging Excel
    prop_path          : path or file-like for the property list Excel
    output_dest        : path or BytesIO to write the output Excel into
    custom_lookup_path : optional path to the vendor-specific custom_lookup JSON
    vendor_name        : optional vendor name (key in VENDOR_CONFIGS) — enables
                         exact column mapping instead of auto-detection
    Returns            : (df_result, review_df, n_total, n_review)
    """
    norm_to_pcode = load_property_list(prop_path)
    custom_lookup = load_custom_lookup(custom_lookup_path) if custom_lookup_path else {}

    df, prop_col, invoice_col = _load_vendor_df(vendor_path, vendor_name)

    # Clean the property column
    df[prop_col] = df[prop_col].astype(str).str.strip()
    df = df[df[prop_col].notna() & (df[prop_col] != "nan") & (df[prop_col] != "")].copy()

    # For Sherwin-Williams, drop the summary "Total" row
    if vendor_name == "Sherwin-Williams":
        df = df[~df[prop_col].str.upper().isin(["TOTAL", "GRAND TOTAL"])].copy()

    # For HD Supply, the parent company row appears with the company name — skip it
    if vendor_name == "HD Supply":
        df = df[~df[prop_col].str.contains("Monarch Investment", case=False, na=False)].copy()

    results = [match_vendor_name(v, norm_to_pcode, fuzzy_threshold, custom_lookup)
               for v in df[prop_col]]

    df["Matched Pcode"]         = [r["pcode"]         for r in results]
    df["Matched Property Name"] = [r["official_name"] for r in results]
    df["Match Confidence"]      = [r["confidence"]    for r in results]
    df["Needs Review"]          = [r["needs_review"]  for r in results]
    df["Match Method"]          = [r["match_method"]  for r in results]

    n_total  = len(df)
    n_review = int(df["Needs Review"].sum())

    # review_df: ALL flagged rows — includes every invoice row
    review_keep = ([invoice_col] if invoice_col else []) + \
                  [prop_col, "Matched Pcode", "Matched Property Name",
                   "Match Confidence", "Match Method"]
    review_df = (df[df["Needs Review"] == True][review_keep]
                 .reset_index(drop=True))

    _write_excel(df, prop_col, output_dest)
    return df, review_df, n_total, n_review


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_excel(df, customer_col, dest):
    HEADER_BG = "1F3864"
    HEADER_FG = "FFFFFF"
    REVIEW_BG = "FFF2CC"
    HIGH_BG   = "E2EFDA"
    MED_BG    = "FCE4D6"

    thin = Side(style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont  = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
    dfont  = Font(name="Arial", size=10)

    extra_cols = ["Matched Pcode", "Matched Property Name", "Match Confidence",
                  "Needs Review", "Match Method"]
    orig_cols  = [c for c in df.columns if c not in extra_cols]
    cust_idx   = orig_cols.index(customer_col) if customer_col in orig_cols else 0
    ordered    = (orig_cols[:cust_idx + 1]
                  + ["Matched Pcode", "Matched Property Name", "Match Confidence", "Needs Review"]
                  + orig_cols[cust_idx + 1:])

    wb = openpyxl.Workbook()

    # --- Sheet 1: Matched Report ---
    ws = wb.active
    ws.title = "Matched Report"
    ws.append(ordered)
    for ci, col in enumerate(ordered, 1):
        c = ws.cell(1, ci)
        c.font = hfont
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 30

    for ri, (_, row) in enumerate(df[ordered].iterrows(), 2):
        nr   = row.get("Needs Review", False)
        conf = row.get("Match Confidence", "HIGH")
        for ci, col in enumerate(ordered, 1):
            val = row[col]
            if pd.isna(val):   val = ""
            elif hasattr(val, "date"): val = val.date()
            c = ws.cell(ri, ci, value=val)
            c.font = dfont
            c.border = border
            c.alignment = Alignment(vertical="center")
            if col in ["Matched Pcode", "Matched Property Name", "Match Confidence", "Needs Review"]:
                if nr:
                    c.fill = PatternFill("solid", fgColor=REVIEW_BG)
                elif conf == "HIGH":
                    c.fill = PatternFill("solid", fgColor=HIGH_BG)
                elif conf == "MEDIUM":
                    c.fill = PatternFill("solid", fgColor=MED_BG)
            if col == "Needs Review":
                c.value = "YES" if val else ""
                c.alignment = Alignment(horizontal="center", vertical="center")

    col_w = {customer_col: 42, "Matched Pcode": 14, "Matched Property Name": 36,
             "Match Confidence": 14, "Needs Review": 13, "Invoice #": 18,
             "Invoice ID": 18, "Invoice": 18, "Number": 18, "Transaction Number": 22,
             "Invoice-suf#": 20, "GL Post Date": 16, "PO #": 16,
             "Subtotal": 12, "Tax": 10, "Grand Total": 13}
    for ci, col in enumerate(ordered, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 16)
    ws.freeze_panes = "A2"

    # --- Sheet 2: Needs Review ---
    ws2 = wb.create_sheet("Needs Review")
    rev_cols = [customer_col, "Matched Pcode", "Matched Property Name",
                "Match Confidence", "Match Method"]
    rev_df   = df[df["Needs Review"] == True][rev_cols].drop_duplicates(customer_col).reset_index(drop=True)

    ws2["A1"] = f"⚠  Needs Review — {len(rev_df)} property names require manual verification"
    ws2["A1"].font = Font(name="Arial", bold=True, size=12, color="C00000")
    ws2.merge_cells(f"A1:{get_column_letter(len(rev_cols))}1")
    ws2.row_dimensions[1].height = 24
    ws2.append(rev_cols)
    for ci, col in enumerate(rev_cols, 1):
        c = ws2.cell(2, ci)
        c.font = hfont; c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border
    for ri, (_, row) in enumerate(rev_df.iterrows(), 3):
        for ci, col in enumerate(rev_cols, 1):
            c = ws2.cell(ri, ci, value=row[col])
            c.font = dfont; c.fill = PatternFill("solid", fgColor=REVIEW_BG)
            c.border = border; c.alignment = Alignment(vertical="center")
    for ci, col in enumerate(rev_cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 20)
    ws2.freeze_panes = "A3"

    # --- Sheet 3: Lookup Reference ---
    ws3 = wb.create_sheet("Lookup Reference")
    ref_cols = [customer_col, "Matched Pcode", "Matched Property Name",
                "Match Confidence", "Needs Review", "Match Method"]
    ws3["A1"] = f"Vendor → Pcode Lookup  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws3["A1"].font = Font(name="Arial", bold=True, size=11, color="1F3864")
    ws3.merge_cells(f"A1:{get_column_letter(len(ref_cols))}1")
    ws3.row_dimensions[1].height = 20
    ws3.append(ref_cols)
    for ci, col in enumerate(ref_cols, 1):
        c = ws3.cell(2, ci)
        c.font = hfont; c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border

    unique = (df[ref_cols].drop_duplicates(customer_col)
              .sort_values(customer_col).reset_index(drop=True))
    for ri, (_, row) in enumerate(unique.iterrows(), 3):
        for ci, col in enumerate(ref_cols, 1):
            val = row[col]
            c = ws3.cell(ri, ci, value=val)
            c.font = dfont; c.border = border; c.alignment = Alignment(vertical="center")
            if col == "Needs Review":
                c.value = "YES" if val else ""
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill("solid", fgColor=REVIEW_BG if val else HIGH_BG)
    ws3.column_dimensions["A"].width = 42
    for ci, col in enumerate(ref_cols[1:], 2):
        ws3.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 18)
    ws3.freeze_panes = "A3"

    wb.save(dest)
